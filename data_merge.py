#automation and new changes  added cost related columns only and changed to access from config file 
import pandas as pd
import mysql.connector
from mysql.connector import Error
from typing import List, Dict, Optional
import logging
import os
import time
from pathlib import Path
import glob
import schedule
import threading
from datetime import datetime
import json
import paramiko
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data_splitter import split_by_vendor_k3_amount

# Configure enhanced logging for automated execution
log_file = f"data_merge_{datetime.now().strftime('%Y%m%d')}.log"
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ====================================================================
# CONFIGURATION LOADING
# ====================================================================

def load_config(config_path: str = "config.json") -> Dict:
    """Load configuration from JSON file."""
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        logger.info(f"Configuration loaded from {config_path}")
        return config
    except FileNotFoundError:
        logger.error(f"Configuration file {config_path} not found")
        raise
    except json.JSONDecodeError as e:
        logger.error(f"Error parsing configuration file: {e}")
        raise
    except Exception as e:
        logger.error(f"Error loading configuration: {e}")
        raise

# Load configuration
CONFIG = load_config()

# Extract configuration values
INPUT_DIRECTORY = CONFIG["input_directory"]
OUTPUT_DIRECTORY = CONFIG["output_directory"]
SUPPORTED_EXTENSIONS = CONFIG["supported_extensions"]
DB_CONFIG = CONFIG["database"]
TABLE_NAME = CONFIG["table_name"]
COLUMN_MAPPING = CONFIG["column_mapping"]
POSSIBLE_REFERENCE_COMBINATIONS = CONFIG["possible_reference_combinations"]
BATCH_SIZE = CONFIG["processing"]["batch_size"]
MAX_RETRIES = CONFIG["processing"]["max_retries"]
CONNECTION_TIMEOUT = CONFIG["processing"]["connection_timeout"]
QUERY_TIMEOUT = CONFIG["processing"]["query_timeout"]
DEBUG_MODE = CONFIG["debug"]["debug_mode"]
DEBUG_ID = CONFIG["debug"]["debug_id"]
SFTP_CONFIG = CONFIG.get("sftp", {})
EMAIL_CONFIG = CONFIG.get("email", {})

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)

# ====================================================================


class FileProcessor:
    """Handles file discovery and batch processing for multiple files."""
    
    def __init__(self, input_directory: str, output_directory: str, supported_extensions: List[str]):
        self.input_directory = input_directory
        self.output_directory = output_directory
        self.supported_extensions = supported_extensions
    
    def discover_files(self) -> List[str]:
        """Discover all supported files in the input directory."""
        files = []
        try:
            for ext in self.supported_extensions:
                pattern = os.path.join(self.input_directory, f"*{ext}")
                found_files = glob.glob(pattern)
                files.extend(found_files)
            
            logger.info(f"Found {len(files)} files to process in {self.input_directory}")
            return files
        except Exception as e:
            logger.error(f"Error discovering files: {e}")
            return []
    
    def get_output_path(self, input_file: str) -> str:
        """Generate output path for processed file with unique filename."""
        filename = os.path.basename(input_file)
        name, ext = os.path.splitext(filename)
        
        # Add timestamp with microseconds for better uniqueness
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d%H%M%S%f")  # Format: YYYYMMDDHHMMSSFFFFFF (includes microseconds)
        
        # Determine output extension
        if ext.lower() == '.csv':
            output_ext = '.csv'
        else:
            output_ext = '.xlsx'
        
        # Generate base output path
        output_path = os.path.join(self.output_directory, f"{name}_{timestamp}{output_ext}")
        
        # If file already exists, append a counter to make it unique
        counter = 1
        while os.path.exists(output_path):
            output_path = os.path.join(self.output_directory, f"{name}_{timestamp}_{counter}{output_ext}")
            counter += 1
        
        return output_path
    
    def move_processed_file(self, input_file: str) -> bool:
        """Move processed file to avoid reprocessing."""
        try:
            filename = os.path.basename(input_file)
            processed_dir = os.path.join(self.input_directory, "processed")
            os.makedirs(processed_dir, exist_ok=True)
            
            destination = os.path.join(processed_dir, filename)
            os.rename(input_file, destination)
            logger.info(f"Moved processed file to: {destination}")
            return True
        except Exception as e:
            logger.error(f"Error moving file {input_file}: {e}")
            return False


class DataEnricher:
    """
    Enhanced data enricher with improved error handling, retry logic, and performance optimizations.
    """
    
    def __init__(self, host: str, database: str, user: str, password: str, 
                 port: int = 3306, debug_mode: bool = False, debug_id: Optional[int] = None):
        """Initialize database connection parameters."""
        self.host = host
        self.database = database
        self.user = user
        self.password = password
        self.port = port
        self.connection = None
        self.debug_mode = debug_mode
        self.debug_id = debug_id
        self.connection_attempts = 0
    
    def connect(self) -> bool:
        """Establish connection to MySQL database with retry logic."""
        for attempt in range(MAX_RETRIES):
            try:
                self.connection = mysql.connector.connect(
                    host=self.host,
                    database=self.database,
                    user=self.user,
                    password=self.password,
                    port=self.port,
                    connection_timeout=CONNECTION_TIMEOUT,
                    autocommit=True
                )
                if self.connection.is_connected():
                    logger.info("Connected to MySQL database")
                    self.connection_attempts = 0
                    return True
            except Error as e:
                self.connection_attempts += 1
                logger.warning(f"Connection attempt {attempt + 1} failed: {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                else:
                    logger.error("Failed to connect to database after all retries")
                    return False
        return False
    
    def disconnect(self):
        """Close database connection safely."""
        if self.connection and self.connection.is_connected():
            self.connection.close()
            logger.info("Database connection closed")
    
    def validate_file(self, file_path: str) -> bool:
        """Validate file exists and is readable."""
        try:
            path = Path(file_path)
            if not path.exists():
                logger.error(f"File does not exist: {file_path}")
                return False
            if not path.is_file():
                logger.error(f"Path is not a file: {file_path}")
                return False
            if not os.access(file_path, os.R_OK):
                logger.error(f"File is not readable: {file_path}")
                return False
            return True
        except Exception as e:
            logger.error(f"Error validating file: {e}")
            return False
    
    def read_file_safely(self, file_path: str):
        """
        Read file and return either:
        - pd.DataFrame for single sheet/CSV files
        - Dict[str, pd.DataFrame] for multi-sheet Excel files
        """
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            if file_extension in ['.xlsx', '.xls', '.xlsb']:
                # Determine engine based on file extension
                if file_extension == '.xlsb':
                    engine = 'pyxlsb'
                elif file_extension == '.xls':
                    engine = None  # pandas will use xlrd or openpyxl automatically
                else:
                    engine = 'openpyxl'
                
                # Read all sheets from Excel file
                try:
                    excel_file = pd.ExcelFile(file_path, engine=engine)
                except ImportError as e:
                    if file_extension == '.xlsb':
                        logger.error(f"pyxlsb library is required to read .xlsb files. Install it with: pip install pyxlsb")
                        raise ImportError("pyxlsb library is required to read .xlsb files. Install it with: pip install pyxlsb")
                    raise
                
                sheet_names = excel_file.sheet_names
                logger.info(f"Found {len(sheet_names)} sheet(s): {sheet_names}")
                
                sheets_dict = {}
                for sheet_name in sheet_names:
                    # Detect the correct header row automatically for each sheet
                    preview = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10, header=None, engine=engine)
                    header_row = None
                    for i, row in preview.iterrows():
                        # Heuristic: a row is header if most cells are strings and not NaN
                        non_null = row.dropna()
                        if len(non_null) > 2 and all(isinstance(x, str) for x in non_null):
                            header_row = i
                            break

                    if header_row is not None:
                        df_sheet = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, engine=engine)
                    else:
                        df_sheet = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                    
                    # Remove unnamed columns (columns that start with "Unnamed:")
                    unnamed_cols = [col for col in df_sheet.columns if str(col).startswith('Unnamed:')]
                    if unnamed_cols:
                        df_sheet = df_sheet.drop(columns=unnamed_cols)
                    
                    if len(df_sheet) > 0:
                        sheets_dict[sheet_name] = df_sheet
                        logger.info(f"Loaded sheet '{sheet_name}' with {len(df_sheet)} rows")
                
                # Return dict if multiple sheets, single DataFrame if one sheet
                if len(sheets_dict) == 0:
                    return pd.DataFrame()
                elif len(sheets_dict) == 1:
                    df = list(sheets_dict.values())[0]
                    logger.info(f"File loaded with columns: {list(df.columns)}")
                    return df
                else:
                    logger.info(f"Returning {len(sheets_dict)} separate sheets")
                    return sheets_dict
            else:
                df = pd.read_csv(file_path)
                # Remove unnamed columns (columns that start with "Unnamed:")
                unnamed_cols = [col for col in df.columns if str(col).startswith('Unnamed:')]
                if unnamed_cols:
                    df = df.drop(columns=unnamed_cols)
                    logger.info(f"Removed {len(unnamed_cols)} unnamed columns: {unnamed_cols}")
                logger.info(f"File loaded with columns: {list(df.columns)}")
                return df
        except Exception as e:
            logger.error(f"Error reading file: {e}")
            return None

    def get_all_columns(self, table_name: str) -> List[str]:
        """Get all column names from the database table with retry logic."""
        for attempt in range(MAX_RETRIES):
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
                columns = [column[0] for column in cursor.fetchall()]
                cursor.close()
                return columns
            except Error as e:
                logger.warning(f"Error fetching columns (attempt {attempt + 1}): {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(1)
                else:
                    logger.error("Failed to fetch columns after all retries")
                    return []
    
    def execute_query_with_retry(self, query: str, params: List = None) -> List[Dict]:
        """Execute query with retry logic and timeout."""
        for attempt in range(MAX_RETRIES):
            try:
                cursor = self.connection.cursor(dictionary=True)
                cursor.execute(query, params or [])
                results = cursor.fetchall()
                cursor.close()
                return results
            except Error as e:
                logger.warning(f"Query failed (attempt {attempt + 1}): {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(1)
                    # Try to reconnect if connection is lost
                    if not self.connection.is_connected():
                        logger.info("Reconnecting to database...")
                        self.connect()
                else:
                    logger.error(f"Query failed after all retries: {query}")
                    return []
    
    def is_empty_value(self, value) -> bool:
        """Check if a value is empty/null."""
        return pd.isna(value) or value is None or (isinstance(value, str) and value.strip() == '')
    
    def find_column_case_insensitive(self, column_name: str, excel_columns: List[str]) -> Optional[str]:
        """Find a column name in Excel columns using case-insensitive matching."""
        column_name_lower = str(column_name).lower().strip()
        for excel_col in excel_columns:
            if str(excel_col).lower().strip() == column_name_lower:
                return excel_col
        return None
    
    def debug_log(self, message: str):
        """Log debug messages only if debug mode is enabled."""
        if self.debug_mode:
            logger.info(f"[DEBUG] {message}")
    
    def split_multi_sector(self, sector: str) -> List[str]:
        """
        Split multi-sector route into individual sectors by '/' separator.
        Examples:
            'MAA-HYD' -> ['MAA-HYD'] (no '/', returns as single sector)
            'MAA-HYD/HYD-MAA' -> ['MAA-HYD', 'HYD-MAA'] (split by '/')
            'MAA-HYD/BLR-DEL/MAA-BOM' -> ['MAA-HYD', 'BLR-DEL', 'MAA-BOM'] (split by '/')
        If sector has no '/', returns original sector as single-item list.
        """
        if self.is_empty_value(sector):
            return []
        
        sector_str = str(sector).strip()
        
        # Split by '/' to get individual sectors
        if '/' in sector_str:
            sectors = [s.strip() for s in sector_str.split('/') if s.strip()]
            return sectors
        else:
            # No '/' found, return as single sector
            return [sector_str]
    
    def get_first_last_sector(self, sector: str) -> Optional[str]:
        """
        Get the first-last airport pattern from a multi-sector route.
        This ignores all intermediate airports.
        Examples:
            'MAA-HYD' -> 'MAA-HYD' (2 airports, no change)
            'MAA-HYD-BBI' -> 'MAA-BBI' (ignores HYD)
            'MAA-HYD-BLR-DEL' -> 'MAA-DEL' (ignores HYD, BLR)
            'MAA-HYD-BLR-DEL-BOM' -> 'MAA-BOM' (ignores HYD, BLR, DEL)
        Returns None if sector is invalid or has less than 2 airports.
        """
        if self.is_empty_value(sector):
            return None
        
        sector_str = str(sector).strip()
        airports = sector_str.split('-')
        
        # Need at least 2 airports
        if len(airports) < 2:
            return None
        
        # If only 2 airports, return as is
        if len(airports) == 2:
            return sector_str
        
        # Return first and last airport
        return f"{airports[0]}-{airports[-1]}"
    
    def detect_reference_columns(self, df_excel: pd.DataFrame, 
                                possible_combinations: List[List[str]]) -> List[str]:
        """
        Detect which reference column combination is available in the Excel file (exact match only).
        """
        available_columns = set(df_excel.columns)
        
        for combination in possible_combinations:
            if all(col in available_columns for col in combination):
                logger.info(f"Detected reference columns: {combination}")
                return combination
        
        logger.error("No suitable reference columns found")
        return []
    
    def create_dynamic_query(self, reference_columns: List[str], params: List) -> str:
        """
        Create dynamic WHERE clause based on available reference columns.
        """
        conditions = []
        
        for ref_col in reference_columns:
            conditions.append(f"`{ref_col}` = %s")
        
        return " AND ".join(conditions)
    
    def try_match_with_combination(self, row: pd.Series, combination: List[str], 
                                   excel_to_db_mapping: Dict[str, str], 
                                   table_name: str, missing_columns: List[str]) -> Optional[Dict]:
        """
        Try to match a single row with database using a specific combination.
        Returns matched data dict if found, None otherwise.
        """
        # Get values for this combination
        key_values = []
        for ref_col in combination:
            # Find the Excel column name that maps to this DB column
            excel_col = None
            for excel_name, db_name in excel_to_db_mapping.items():
                if db_name == ref_col:
                    excel_col = excel_name
                    break
            # If not found in mapping, try case-insensitive search (need to get df_excel from context)
            # Note: This method might not have direct access to df_excel, so we'll use a simpler approach
            if excel_col is None:
                # Try to find column in row.index case-insensitively
                for col in row.index:
                    if str(col).lower().strip() == str(ref_col).lower().strip():
                        excel_col = col
                        break
            # Last resort: use ref_col as-is
            if excel_col is None:
                excel_col = ref_col
            
            # Check if column exists and has non-empty value
            if excel_col not in row.index:
                return None  # Column not available
            
            value = row[excel_col]
            if self.is_empty_value(value):
                return None  # Empty value, can't use this combination
            
            key_values.append(value)
        
        # Build and execute query
        ref_cols_str = ', '.join([f"`{c}`" for c in combination])
        missing_columns_str = ', '.join([f"`{col}`" for col in missing_columns])
        
        # Build WHERE clause with AND conditions
        where_conditions = ' AND '.join([f"`{c}` = %s" for c in combination])
        
        query = (
            f"SELECT {ref_cols_str}, {missing_columns_str} "
            f"FROM `{table_name}` "
            f"WHERE {where_conditions} "
            f"LIMIT 1"
        )
        
        results = self.execute_query_with_retry(query, key_values)
        
        if results and len(results) > 0:
            return {col: results[0].get(col) for col in missing_columns}
        
        return None
    
    def apply_header_formatting(self, original_excel_path: str, output_excel_path: str, 
                                header_row_index: int = 1) -> bool:
        """
        Apply header row formatting, column widths, and row heights from original Excel file to output file.
        Optimized to only modify header row and dimensions, preserving all other formatting.
        Handles multiple sheets by matching sheet names.
        
        Args:
            original_excel_path: Path to original Excel file
            output_excel_path: Path to output Excel file
            header_row_index: Row index for header (1-based, default 1)
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Skip formatting for .xlsb files as openpyxl doesn't support them
            original_ext = os.path.splitext(original_excel_path)[1].lower()
            if original_ext == '.xlsb':
                logger.info("Skipping header formatting for .xlsb file (openpyxl doesn't support .xlsb format)")
                return False
            
            # Load original file to get header formatting
            original_wb = load_workbook(original_excel_path, read_only=False, data_only=False)
            # Load output file for writing
            output_wb = load_workbook(output_excel_path)
            
            # Process each sheet in output file
            for sheet_name in output_wb.sheetnames:
                if sheet_name in original_wb.sheetnames:
                    original_ws = original_wb[sheet_name]
                    output_ws = output_wb[sheet_name]
                    self._apply_sheet_formatting(original_ws, output_ws, header_row_index)
                else:
                    # If sheet doesn't exist in original, use first sheet as template
                    original_ws = original_wb.active
                    output_ws = output_wb[sheet_name]
                    self._apply_sheet_formatting(original_ws, output_ws, header_row_index)
            
            # Save the formatted output file
            output_wb.save(output_excel_path)
            output_wb.close()
            original_wb.close()
            
            logger.info(f"Applied header formatting from original file to output")
            return True
            
        except Exception as e:
            logger.warning(f"Could not apply header formatting: {e}")
            try:
                if 'output_wb' in locals():
                    output_wb.close()
                if 'original_wb' in locals():
                    original_wb.close()
            except Exception:
                pass
            return False
    
    def _apply_sheet_formatting(self, original_ws, output_ws, header_row_index: int):
        """Helper method to apply formatting to a single sheet."""
        # Find header row in original file (search first few rows)
        original_header_row = None
        for row_idx in range(1, min(11, original_ws.max_row + 1)):
            row = original_ws[row_idx]
            # Check if this row has mostly text values (likely header)
            text_count = sum(1 for cell in row if cell.value and isinstance(cell.value, str))
            if text_count >= len(row) * 0.5:  # At least 50% text
                original_header_row = row_idx
                break
        
        if original_header_row is None:
            original_header_row = 1  # Default to first row
        
        # Copy column widths from original to output
        # Strategy: Match by header name first, then by position for existing columns
        # For new columns at the end, use average width or default
        
        # Map original columns by header name and position
        original_headers_by_name = {}
        original_widths_by_pos = {}
        
        for col_idx in range(1, original_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            cell = original_ws.cell(row=original_header_row, column=col_idx)
            if cell.value:
                header_name = str(cell.value).strip()
                original_headers_by_name[header_name] = col_idx
            
            # Store width by position (even if no header)
            if col_letter in original_ws.column_dimensions:
                width = original_ws.column_dimensions[col_letter].width
                if width and width > 0:
                    original_widths_by_pos[col_idx] = width
        
        # Map output columns by header name
        output_headers_by_name = {}
        for col_idx in range(1, output_ws.max_column + 1):
            cell = output_ws.cell(row=header_row_index, column=col_idx)
            if cell.value:
                header_name = str(cell.value).strip()
                output_headers_by_name[header_name] = col_idx
        
        # Calculate default width (average of original widths, or 12.0)
        widths_list = [w for w in original_widths_by_pos.values() if w > 0]
        default_width = sum(widths_list) / len(widths_list) if widths_list else 12.0
        
        # Copy widths: first match by header name, then by position
        for out_col_idx in range(1, output_ws.max_column + 1):
            out_col_letter = get_column_letter(out_col_idx)
            out_cell = output_ws.cell(row=header_row_index, column=out_col_idx)
            out_header_name = str(out_cell.value).strip() if out_cell.value else None
            
            width_to_apply = None
            
            # Try to match by header name first
            if out_header_name and out_header_name in original_headers_by_name:
                orig_col_idx = original_headers_by_name[out_header_name]
                orig_col_letter = get_column_letter(orig_col_idx)
                if orig_col_letter in original_ws.column_dimensions:
                    width_to_apply = original_ws.column_dimensions[orig_col_letter].width
            
            # If no match by name, try by position (for existing columns)
            if (width_to_apply is None or width_to_apply == 0) and out_col_idx <= original_ws.max_column:
                width_to_apply = original_widths_by_pos.get(out_col_idx)
            
            # If still no width, use default (for new columns)
            if width_to_apply is None or width_to_apply == 0:
                width_to_apply = default_width
            
            # Apply the width
            output_ws.column_dimensions[out_col_letter].width = width_to_apply
        
        # Copy header row height from original file
        if original_header_row in original_ws.row_dimensions:
            orig_row_height = original_ws.row_dimensions[original_header_row].height
            if orig_row_height and orig_row_height > 0:
                output_ws.row_dimensions[header_row_index].height = orig_row_height
        
        # Get formatting from first existing header cell in original file
        sample_cell = None
        for col_idx in range(1, original_ws.max_column + 1):
            cell = original_ws.cell(row=original_header_row, column=col_idx)
            if cell.value:  # Find first non-empty cell
                sample_cell = cell
                break
        
        if sample_cell is None:
            # If no sample found, try to get any cell from header row
            if original_ws.max_column > 0:
                sample_cell = original_ws.cell(row=original_header_row, column=1)
        
        # Apply formatting to all header cells in output file
        if sample_cell:
            # Copy style properties from sample cell
            header_fill = None
            if sample_cell.fill and hasattr(sample_cell.fill, 'start_color'):
                fill_color = sample_cell.fill.start_color
                # Check if fill has a meaningful color (not default/transparent)
                if fill_color and (fill_color.index not in [None, '00000000', 'FFFFFFFF'] or 
                                  (hasattr(fill_color, 'rgb') and fill_color.rgb)):
                    header_fill = sample_cell.fill
            
            header_font = sample_cell.font if sample_cell.font else None
            header_alignment = sample_cell.alignment if sample_cell.alignment else None
            header_border = sample_cell.border if sample_cell.border else None
            
            # Apply to all columns in output header row (including new columns)
            for col_idx in range(1, output_ws.max_column + 1):
                cell = output_ws.cell(row=header_row_index, column=col_idx)
                
                # Apply fill (background color)
                if header_fill:
                    try:
                        # Try to copy the fill object directly (more reliable)
                        if hasattr(header_fill, 'copy'):
                            cell.fill = header_fill.copy()
                        else:
                            cell.fill = PatternFill(
                                fill_type=header_fill.fill_type,
                                start_color=header_fill.start_color,
                                end_color=header_fill.end_color
                            )
                    except Exception as fill_err:
                        logger.debug(f"Could not copy fill style: {fill_err}")
                        pass
                
                # Apply font
                if header_font:
                    try:
                        cell.font = Font(
                            name=header_font.name or 'Calibri',
                            size=header_font.size or 11,
                            bold=header_font.bold,
                            italic=header_font.italic,
                            underline=header_font.underline,
                            strike=header_font.strike,
                            color=header_font.color
                        )
                    except Exception:
                        pass
                
                # Apply alignment
                if header_alignment:
                    try:
                        cell.alignment = Alignment(
                            horizontal=header_alignment.horizontal or 'general',
                            vertical=header_alignment.vertical or 'bottom',
                            wrap_text=header_alignment.wrap_text,
                            shrink_to_fit=header_alignment.shrink_to_fit,
                            indent=header_alignment.indent
                        )
                    except Exception:
                        pass
                
                # Apply border
                if header_border:
                    try:
                        cell.border = Border(
                            left=header_border.left,
                            right=header_border.right,
                            top=header_border.top,
                            bottom=header_border.bottom
                        )
                    except Exception:
                        pass
    
    def format_date_columns(self, output_excel_path: str, header_row_index: int = 1) -> bool:
        """
        Format date columns in the output Excel file to display as dd-mm-yyyy.
        
        Args:
            output_excel_path: Path to output Excel file
            header_row_index: Row index for header (1-based, default 1)
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # List of date column names to format (exact matches and partial matches)
            date_columns = [
                'Booking_Date', 'Booking Date', 'booking date',
                'Travel_Date', 'Travel Date', 'travel date', 'Departure Date', 'departure date',
                'Invoice_Received_Date', 'Invoice Received Date', 'invoice received date',
                'Invoice_Date', 'Invoice Date', 'invoice date', 'Date Of Invoice', 'date of invoice',
                'Original_Invoice_Date', 'Original Invoice Date', 'original invoice date'
            ]
            
            # Also check for columns containing "date" keyword
            date_keywords = ['date', 'Date', 'DATE']
            
            # Load output file
            output_wb = load_workbook(output_excel_path)
            
            # Process each sheet
            for sheet_name in output_wb.sheetnames:
                output_ws = output_wb[sheet_name]
                
                # Find date columns by header name
                date_col_indices = []
                for col_idx in range(1, output_ws.max_column + 1):
                    cell = output_ws.cell(row=header_row_index, column=col_idx)
                    if cell.value:
                        header_name = str(cell.value).strip()
                        header_lower = header_name.lower()
                        
                        # Check if this header matches any date column (case-insensitive)
                        is_date_column = False
                        for date_col in date_columns:
                            if date_col.lower() in header_lower or header_lower in date_col.lower():
                                is_date_column = True
                                break
                        
                        # Also check if header contains "date" keyword (but not "update", "validate", etc.)
                        if not is_date_column:
                            if 'date' in header_lower and not any(word in header_lower for word in ['update', 'validate', 'created', 'modified']):
                                is_date_column = True
                        
                        if is_date_column:
                            date_col_indices.append(col_idx)
                            logger.info(f"Found date column: {header_name} at column {col_idx}")
                
                # Apply date formatting to all rows in date columns
                date_format = 'dd-mm-yyyy'
                for col_idx in date_col_indices:
                    col_letter = get_column_letter(col_idx)
                    formatted_count = 0
                    # Format all data rows (skip header row)
                    for row_idx in range(header_row_index + 1, output_ws.max_row + 1):
                        cell = output_ws.cell(row=row_idx, column=col_idx)
                        # Only format if cell has a value
                        if cell.value is not None:
                            # Apply date format to all non-empty cells in date columns
                            # This will format both Excel date serials and datetime objects
                            cell.number_format = date_format
                            formatted_count += 1
                    
                    if formatted_count > 0:
                        logger.info(f"Formatted {formatted_count} date cells in column {col_letter}")
            
            # Save the formatted file
            output_wb.save(output_excel_path)
            output_wb.close()
            
            logger.info(f"Applied date formatting to output file")
            return True
            
        except Exception as e:
            logger.warning(f"Could not apply date formatting: {e}")
            try:
                if 'output_wb' in locals():
                    output_wb.close()
            except Exception:
                pass
            return False
    
    def _enrich_single_dataframe(self, df_excel: pd.DataFrame, table_name: str,
                                 possible_reference_combinations: List[List[str]],
                                 column_mapping: Dict[str, str]) -> Optional[pd.DataFrame]:
        """Helper method to enrich a single DataFrame."""
        logger.info(f"Processing {len(df_excel)} rows...")
        logger.info(f"Available columns: {list(df_excel.columns)}")
        
        # Create column mapping for database operations (without renaming Excel columns)
        excel_to_db_mapping = {}
        if column_mapping:
            excel_columns_list = list(df_excel.columns)
            for mapping_key, db_col in column_mapping.items():
                if mapping_key in df_excel.columns:
                    excel_to_db_mapping[mapping_key] = db_col
                else:
                    matched_col = self.find_column_case_insensitive(mapping_key, excel_columns_list)
                    if matched_col:
                        excel_to_db_mapping[matched_col] = db_col
                        logger.info(f"Matched '{mapping_key}' (from config) to '{matched_col}' (in Excel) - case-insensitive match")
            logger.info(f"Created mapping for {len(excel_to_db_mapping)} columns")
        
        # Create a temporary DataFrame with mapped column names for reference detection
        df_temp = df_excel.copy()
        if excel_to_db_mapping:
            df_temp = df_temp.rename(columns=excel_to_db_mapping)
        
        # Check if at least one combination is available (for validation)
        available_columns = set(df_temp.columns)
        valid_combinations = []
        for combination in possible_reference_combinations:
            if all(col in available_columns for col in combination):
                valid_combinations.append(combination)
        
        if not valid_combinations:
            logger.error("No suitable reference columns found")
            return None
        
        logger.info(f"Will try matching with {len(valid_combinations)} combinations in cascade order: {valid_combinations}")
        
        # Get database columns
        all_db_columns = self.get_all_columns(table_name)
        if not all_db_columns:
            logger.error("Failed to get database columns")
            return None
        
        # Validate that all columns in combinations exist in the database
        all_db_columns_set = set(all_db_columns)
        validated_combinations = []
        for combination in valid_combinations:
            if all(col in all_db_columns_set for col in combination):
                validated_combinations.append(combination)
            else:
                missing_cols = [col for col in combination if col not in all_db_columns_set]
                logger.warning(f"Skipping combination {combination} - database columns not found: {missing_cols}")
        
        if not validated_combinations:
            logger.error("No valid combinations found after database column validation")
            return None
        
        # Update valid_combinations to only include validated ones
        valid_combinations = validated_combinations
        logger.info(f"After database validation, {len(valid_combinations)} combinations are valid: {valid_combinations}")
        
        # Define output column structure: (db_column_name, output_column_name, excel_only)
        # excel_only=True means only take from Excel if column exists, not from DB
        output_column_mapping = [
            (None, 'Legal_Name', True),  # Excel only: "GST Name"
            (None, 'Company_GST_Number', True),  # Excel only: "GST Number"
            (None, 'Booking_Date', True),  # Excel only: "Booking Date"
            (None, 'Travel_Date', True),  # Excel only: "Departure Date"
            (None, 'Passenger_Name', True),  # Excel only: "LOUIS ARUL AROCKIASAMY" or similar passenger name columns
            (None, 'PNR', True),  # Excel only: "Airline Pnr"
            (None, 'Ticket_Number', True),  # Excel only: "Airline Ticket No."
            (None, 'Vendor Invoice No', True),  # Excel only: "GST INVOICE NO"
            (None, 'Vendor K3 Amount', True),  # Excel only: "TOTAL K3"
            (None, 'Airline_Name', True),  # Excel only: "Airline Name"
            (None, 'Airline_Code', True),  # Excel only: "Airline Code"
            (None, 'Travel_Mode', True),  # Excel only: "TRIP TYPE"
            (None, 'Travel_Sector', True),  # Excel only: "Sector"
            ('Place_Of_Supply', 'Embarking_State', False),  # From DB
            ('Airline_Gst_Number', 'Airline_GST_Number', False),  # From DB
            ('Airline_Gst_Name', 'Airline_Legal_Name', False),  # From DB
            (None, 'Ticket_Amount', True),  # Excel only: "Total Fare (Including GST)"
            (None, 'Cost_Center', True),  # Excel only: "Cost_Center"
            ('Invoice_Type', 'Invoice_Type', False),  # From DB
            ('Email_Date', 'Invoice_Received_Date', False),  # From DB
            ('Date_Of_Invoice', 'Invoice_Date', False),  # From DB
            ('Invoice_Number', 'Invoice_Number', False),  # From DB
            ('Original_Invoice_Number', 'Original_Invoice_Number', False),  # From DB
            ('Original_Invoice_Date', 'Original_Invoice_Date', False),  # From DB
            ('Invoice_Total', 'Total_Invoice_Amount', False),  # From DB
            ('Taxable_Amount', 'Invoice_Taxable_Value', False),  # From DB
            ('NonTaxable_Amount', 'Invoice_Non_Taxable_Value', False),  # From DB
            ('Igst_Total', 'Invoice_IGST', False),  # From DB
            ('Cgst_Total', 'Invoice_CGST', False),  # From DB
            ('Sgst_Total', 'Invoice_SGST', False),  # From DB
            ('Invoice_Total_GST', 'Invoice_Total_GST_Amount', False),  # From DB
            ('Igst_Rate', 'IGST_Rate', False),  # From DB
            ('Cgst_Rate', 'CGST_Rate', False),  # From DB
            ('Sgst_Rate', 'SGST_Rate', False),  # From DB
            ('Public_File_URL', 'Airline_Invoice_Download_Url', False),  # From DB
        ]
        
        # Excel column name mappings for Excel-only columns
        excel_column_mappings = {
            'Legal_Name': ['GST Name', 'gst name', 'GSTName', 'GST_Name', 'gst_name'],
            'Company_GST_Number': ['GST Number', 'gst number', 'GSTNumber', 'GST_Number', 'gst_number'],
            'Booking_Date': ['Booking Date', 'booking date', 'BookingDate', 'Booking_Date', 'booking_date'],
            'Travel_Date': ['Departure Date', 'departure date', 'DepartureDate', 'Departure_Date', 'departure_date', 'Onward Date', 'onward date', 'OnwardDate', 'Travel_Date', 'travel date'],
            'Passenger_Name': ['LOUIS ARUL AROCKIASAMY', 'louis arul arokiasamy', 'Traveller', 'Traveler_Name', 'traveler name', 'TravelerName', 'Passenger_Name', 'passenger name', 'Passenger Name'],
            'PNR': ['Airline Pnr', 'airline pnr', 'Airline PNR', 'Airline PNR/Prov. Booking', 'airline pnr/prov. booking', 'pnrnumber', 'pnr number', 'PNR_Number', 'PNR', 'pnr'],
            'Ticket_Number': ['Airline Ticket No.', 'airline ticket no.', 'Airline Ticket No', 'Ticket Num/Final Booking', 'ticket num/final booking', 'TicketNumber', 'Ticket Number', 'ticket number', 'Ticket_Number'],
            'Vendor Invoice No': ['GST INVOICE NO', 'gst invoice no', 'GST Invoice No', 'GST_INVOICE_NO', 'GSTInvoiceNo', 'GST Invoice Number'],
            'Vendor K3 Amount': ['TOTAL K3', 'total k3', 'Total K3', 'TOTAL_K3', 'TotalK3', 'Total K3 Amount'],
            'Airline_Name': ['Airline Name', 'airline name', 'AirlineName', 'Airline_Name', 'airline_name'],
            'Airline_Code': ['Airline Code', 'airline code', 'AirlineCode', 'airlinecode', 'Airline_Code'],
            'Travel_Mode': ['TRIP TYPE', 'trip type', 'Trip Type', 'TRIP_TYPE', 'TripType', 'Product Type', 'product type', 'ProductType', 'Travel_Mode', 'travel mode'],
            'Travel_Sector': ['Sector', 'sector', 'Travel_Sector', 'travel sector', 'TravelSector', 'trvael sector', 'travelsector'],
            'Ticket_Amount': ['Total Fare (Including GST)', 'total fare (including gst)', 'Total Fare', 'total fare', 'TotalFare', 'Total_Fare'],
            'Cost_Center': ['Cost_Center', 'cost center', 'CostCenter', 'Cost Centre', 'cost centre', 'Cost Center']
        }
        
        # Extract DB columns to fetch (exclude Excel-only columns)
        db_columns_to_fetch = [col[0] for col in output_column_mapping if col[0] is not None and not col[2]]
        # Remove duplicates while preserving order
        db_columns_to_fetch = list(dict.fromkeys(db_columns_to_fetch))
        
        # Numeric columns that should be aggregated when multiple sectors are present
        numeric_columns_to_aggregate = [
            'Taxable_Amount', 'NonTaxable_Amount', 'Cgst_Total', 'Sgst_Total', 
            'Igst_Total', 'Invoice_Total_GST', 'Invoice_Total', 'Igst_Rate', 
            'Cgst_Rate', 'Sgst_Rate'
        ]
        
        # We need to fetch ALL DB columns from database (even if they exist in Excel)
        # User requirement: "initially fetch from db"
        missing_columns = db_columns_to_fetch.copy()
        
        logger.info(f"Will fetch {len(missing_columns)} columns from database: {missing_columns}")
        
        # Find the sector column name in Excel (could be mapped or original)
        sector_excel_col = None
        for excel_name, db_name in excel_to_db_mapping.items():
            if db_name == 'Travel_Sector':
                sector_excel_col = excel_name
                break
        if sector_excel_col is None:
            # Try to find it directly
            for col in df_excel.columns:
                if col.lower().strip() in ['sector', 'travel_sector', 'travel sector']:
                    sector_excel_col = col
                    break
        
        # Process data in batches with cascading matching logic (optimized for performance)
        enriched_data = []
        match_count = 0
        no_match_count = 0
        combination_usage_stats = {str(combo): 0 for combo in valid_combinations}
        
        logger.info(f"Processing {len(df_excel)} rows with cascading matching (batch mode)...")
        
        # Process in batches for better performance
        for batch_start in range(0, len(df_excel), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(df_excel))
            batch_df = df_excel.iloc[batch_start:batch_end]
            
            logger.info(f"Processing batch {batch_start//BATCH_SIZE + 1}: rows {batch_start + 1}-{batch_end}")
            
            # Track which rows have been matched and which combination was used
            row_matches = {}  # idx -> {matched_data, combination}
            unmatched_indices = set(batch_df.index)
            
            # Try each combination in cascade order, only for unmatched rows
            for combination in valid_combinations:
                if not unmatched_indices:
                    break  # All rows matched, skip remaining combinations
                
                # Check if Travel_Sector is in this combination
                has_sector = 'Travel_Sector' in combination
                
                # Build keys for rows that can use this combination
                row_keys = {}  # idx -> list of tuples (for multi-sector, multiple keys per row)
                rows_with_valid_keys = []
                
                for idx in list(unmatched_indices):
                    row = batch_df.loc[idx]
                    key_values_list = []  # List of key tuples (one per sector if multi-sector)
                    valid = True
                    
                    # Get sector value if Travel_Sector is in combination
                    sector_value = None
                    if has_sector and sector_excel_col and sector_excel_col in row.index:
                        sector_value = row[sector_excel_col]
                    
                    # Split sector if it's multi-sector
                    sectors_to_query = []
                    if has_sector and sector_value and not self.is_empty_value(sector_value):
                        sectors_to_query = self.split_multi_sector(sector_value)
                    else:
                        sectors_to_query = [None]  # Single query without sector splitting
                    
                    # Build keys for each sector
                    for sector in sectors_to_query:
                        key_values = []
                        valid_for_sector = True
                        
                        for ref_col in combination:
                            # Find the Excel column name that maps to this DB column
                            excel_col = None
                            for excel_name, db_name in excel_to_db_mapping.items():
                                if db_name == ref_col:
                                    excel_col = excel_name
                                    break
                            # If not found in mapping, try case-insensitive search in original Excel columns
                            if excel_col is None:
                                excel_col = self.find_column_case_insensitive(ref_col, list(df_excel.columns))
                                # If still not found, try to find by checking column_mapping keys
                                if excel_col is None:
                                    for mapping_key, db_name in column_mapping.items():
                                        if db_name == ref_col:
                                            excel_col = self.find_column_case_insensitive(mapping_key, list(df_excel.columns))
                                            if excel_col:
                                                break
                            # Last resort: use ref_col as-is (might work if column name matches exactly)
                            if excel_col is None:
                                excel_col = ref_col
                            
                            # Check if column exists and has non-empty value
                            if excel_col not in row.index:
                                valid_for_sector = False
                                break
                            
                            value = row[excel_col]
                            
                            # If this is Travel_Sector and we have a split sector, use the split sector
                            if ref_col == 'Travel_Sector' and sector is not None:
                                value = sector
                            
                            if self.is_empty_value(value):
                                valid_for_sector = False
                                break
                            
                            key_values.append(value)
                        
                        if valid_for_sector:
                            key_values_list.append(tuple(key_values))
                    
                    if key_values_list:
                        row_keys[idx] = key_values_list
                        rows_with_valid_keys.append(idx)
                
                # If we have valid keys, execute batch query
                if row_keys:
                    # Collect all unique keys (flatten the list of lists)
                    all_keys = []
                    key_to_row_indices = {}  # Map key -> list of row indices that use this key
                    for idx, keys_list in row_keys.items():
                        for key in keys_list:
                            if key not in key_to_row_indices:
                                all_keys.append(key)
                                key_to_row_indices[key] = []
                            key_to_row_indices[key].append(idx)
                    
                    unique_keys = list(dict.fromkeys(all_keys))  # Preserve order, remove duplicates
                    
                    # Build batch query
                    ref_cols_str = ', '.join([f"`{c}`" for c in combination])
                    missing_columns_str = ', '.join([f"`{col}`" for col in missing_columns])
                    placeholders = ', '.join(["(" + ", ".join(["%s"] * len(combination)) + ")" for _ in unique_keys])
                    
                    # Add NULL checks to ensure we don't match on NULL values in database
                    null_checks = ' AND '.join([f"`{c}` IS NOT NULL" for c in combination])
                    
                    query = (
                        f"SELECT {ref_cols_str}, {missing_columns_str} "
                        f"FROM `{table_name}` "
                        f"WHERE ({ref_cols_str}) IN ({placeholders}) "
                        f"AND {null_checks}"
                    )
                    params = [v for key in unique_keys for v in key]
                    
                    # Log query details for Ticket_Number combination
                    if 'Ticket_Number' in combination:
                        logger.info(f"Executing query for combination {combination}: {len(unique_keys)} unique keys")
                        logger.debug(f"Query: {query[:200]}... (truncated)")
                        logger.debug(f"Sample params (first 3): {params[:min(6, len(params))]}")
                    
                    # Execute batch query
                    results = self.execute_query_with_retry(query, params)
                    
                    # Log results for Ticket_Number combination
                    if 'Ticket_Number' in combination:
                        logger.info(f"Query returned {len(results)} results for combination {combination}")
                    
                    # Build lookup dictionary from results - only store first result per key (LIMIT 1 per combination)
                    lookup = {}
                    for r in results:
                        key = tuple(r[c] for c in combination)
                        # Only store the first result for each key (ignore duplicates)
                        if key not in lookup:
                            lookup[key] = {col: r.get(col) for col in missing_columns}
                    
                    # Match rows with results and aggregate for multi-sector
                    for idx in rows_with_valid_keys:
                        if idx in unmatched_indices:  # Only process if not already matched
                            keys_list = row_keys[idx]
                            aggregated_data = {}
                            
                            # Collect all matching results for this row (one per sector/key)
                            all_results = []
                            for key in keys_list:
                                if key in lookup:
                                    # Each key contributes one result (already limited to 1 per key)
                                    all_results.append(lookup[key])
                            
                            if all_results:
                                # Aggregate numeric columns across all sectors
                                for col in missing_columns:
                                    if col in numeric_columns_to_aggregate:
                                        # Sum numeric values from all sectors
                                        total = 0
                                        for result in all_results:
                                            val = result.get(col)
                                            if val is not None:
                                                try:
                                                    total += float(val)
                                                except (ValueError, TypeError):
                                                    pass
                                        aggregated_data[col] = total if total != 0 else None
                                    else:
                                        # For non-numeric columns, take first non-null value
                                        for result in all_results:
                                            val = result.get(col)
                                            if val is not None:
                                                aggregated_data[col] = val
                                                break
                                        if col not in aggregated_data:
                                            aggregated_data[col] = None
                                
                                row_matches[idx] = {
                                    'matched_data': aggregated_data,
                                    'combination': combination
                                }
                                unmatched_indices.remove(idx)
                                combination_usage_stats[str(combination)] += 1
            
            # Fallback: Try first-last sector pattern for unmatched rows
            # This matches patterns like "MAA-HYD-BBI" with "MAA-BBI" in the database
            if unmatched_indices and sector_excel_col:
                logger.info(f"Trying first-last sector pattern fallback for {len(unmatched_indices)} unmatched rows")
                
                # Find combinations that include Travel_Sector
                sector_combinations = [combo for combo in valid_combinations if 'Travel_Sector' in combo]
                
                for combination in sector_combinations:
                    if not unmatched_indices:
                        break
                    
                    # Build keys using first-last sector pattern
                    row_keys_fallback = {}
                    rows_with_valid_keys_fallback = []
                    
                    for idx in list(unmatched_indices):
                        row = batch_df.loc[idx]
                        
                        # Get sector value
                        if sector_excel_col not in row.index:
                            continue
                        
                        sector_value = row[sector_excel_col]
                        if self.is_empty_value(sector_value):
                            continue
                        
                        # Get first-last sector pattern
                        first_last_sector = self.get_first_last_sector(sector_value)
                        if not first_last_sector or first_last_sector == sector_value:
                            # No intermediate airports to skip, skip this fallback
                            continue
                        
                        # Build key with first-last sector
                        key_values = []
                        valid_for_sector = True
                        
                        for ref_col in combination:
                            # Find the Excel column name that maps to this DB column
                            excel_col = None
                            for excel_name, db_name in excel_to_db_mapping.items():
                                if db_name == ref_col:
                                    excel_col = excel_name
                                    break
                            # If not found in mapping, try case-insensitive search in original Excel columns
                            if excel_col is None:
                                excel_col = self.find_column_case_insensitive(ref_col, list(df_excel.columns))
                                # If still not found, try to find by checking column_mapping keys
                                if excel_col is None:
                                    for mapping_key, db_name in column_mapping.items():
                                        if db_name == ref_col:
                                            excel_col = self.find_column_case_insensitive(mapping_key, list(df_excel.columns))
                                            if excel_col:
                                                break
                            # Last resort: use ref_col as-is (might work if column name matches exactly)
                            if excel_col is None:
                                excel_col = ref_col
                            
                            # Check if column exists and has non-empty value
                            if excel_col not in row.index:
                                valid_for_sector = False
                                break
                            
                            value = row[excel_col]
                            
                            # Use first-last sector pattern for Travel_Sector
                            if ref_col == 'Travel_Sector':
                                value = first_last_sector
                            
                            if self.is_empty_value(value):
                                valid_for_sector = False
                                break
                            
                            key_values.append(value)
                        
                        if valid_for_sector:
                            row_keys_fallback[idx] = [tuple(key_values)]
                            rows_with_valid_keys_fallback.append(idx)
                    
                    # If we have valid keys, execute batch query
                    if row_keys_fallback:
                        # Collect all unique keys
                        all_keys_fallback = []
                        key_to_row_indices_fallback = {}
                        for idx, keys_list in row_keys_fallback.items():
                            for key in keys_list:
                                if key not in key_to_row_indices_fallback:
                                    all_keys_fallback.append(key)
                                    key_to_row_indices_fallback[key] = []
                                key_to_row_indices_fallback[key].append(idx)
                        
                        unique_keys_fallback = list(dict.fromkeys(all_keys_fallback))
                        
                        # Build batch query
                        ref_cols_str = ', '.join([f"`{c}`" for c in combination])
                        missing_columns_str = ', '.join([f"`{col}`" for col in missing_columns])
                        placeholders = ', '.join(["(" + ", ".join(["%s"] * len(combination)) + ")" for _ in unique_keys_fallback])
                        
                        # Add NULL checks
                        null_checks = ' AND '.join([f"`{c}` IS NOT NULL" for c in combination])
                        
                        query = (
                            f"SELECT {ref_cols_str}, {missing_columns_str} "
                            f"FROM `{table_name}` "
                            f"WHERE ({ref_cols_str}) IN ({placeholders}) "
                            f"AND {null_checks}"
                        )
                        params = [v for key in unique_keys_fallback for v in key]
                        
                        # Execute batch query
                        results_fallback = self.execute_query_with_retry(query, params)
                        
                        # Build lookup dictionary from results
                        lookup_fallback = {}
                        for r in results_fallback:
                            key = tuple(r[c] for c in combination)
                            if key not in lookup_fallback:
                                lookup_fallback[key] = {col: r.get(col) for col in missing_columns}
                        
                        # Match rows with results
                        for idx in rows_with_valid_keys_fallback:
                            if idx in unmatched_indices:
                                keys_list = row_keys_fallback[idx]
                                
                                # Collect matching results
                                all_results_fallback = []
                                for key in keys_list:
                                    if key in lookup_fallback:
                                        all_results_fallback.append(lookup_fallback[key])
                                
                                if all_results_fallback:
                                    # Aggregate numeric columns
                                    aggregated_data_fallback = {}
                                    for col in missing_columns:
                                        if col in numeric_columns_to_aggregate:
                                            total = 0
                                            for result in all_results_fallback:
                                                val = result.get(col)
                                                if val is not None:
                                                    try:
                                                        total += float(val)
                                                    except (ValueError, TypeError):
                                                        pass
                                            aggregated_data_fallback[col] = total if total != 0 else None
                                        else:
                                            # For non-numeric columns, take first non-null value
                                            for result in all_results_fallback:
                                                val = result.get(col)
                                                if val is not None:
                                                    aggregated_data_fallback[col] = val
                                                    break
                                            if col not in aggregated_data_fallback:
                                                aggregated_data_fallback[col] = None
                                    
                                    row_matches[idx] = {
                                        'matched_data': aggregated_data_fallback,
                                        'combination': combination
                                    }
                                    unmatched_indices.remove(idx)
                                    combination_usage_stats[f"{str(combination)} (first-last fallback)"] = combination_usage_stats.get(f"{str(combination)} (first-last fallback)", 0) + 1
                                    logger.debug(f"Matched row {idx} using first-last sector pattern: {first_last_sector}")
            
            # Fallback: Try Ticket_Number value as PNR_Number (since Ticket_Number column sometimes contains PNR numbers)
            # This handles cases where Ticket_Number column has PNR values instead of ticket numbers
            if unmatched_indices and sector_excel_col:
                logger.info(f"Trying Ticket_Number as PNR_Number fallback for {len(unmatched_indices)} unmatched rows")
                
                # Find Ticket_Number column in Excel
                ticket_excel_col = None
                for excel_name, db_name in excel_to_db_mapping.items():
                    if db_name == 'Ticket_Number':
                        ticket_excel_col = excel_name
                        break
                if ticket_excel_col is None:
                    # Try to find it directly
                    for col in df_excel.columns:
                        if col.lower().strip() in ['ticket number', 'ticketnumber', 'ticket_num', 'airline ticket no.', 'ticket num/final booking']:
                            ticket_excel_col = col
                            break
                
                # Find PNR_Number column in Excel (for reference, but we'll use Ticket_Number value)
                pnr_excel_col = None
                for excel_name, db_name in excel_to_db_mapping.items():
                    if db_name == 'PNR_Number':
                        pnr_excel_col = excel_name
                        break
                if pnr_excel_col is None:
                    for col in df_excel.columns:
                        if col.lower().strip() in ['pnr', 'pnr number', 'pnrnumber', 'airline pnr', 'airline pnr/prov. booking']:
                            pnr_excel_col = col
                            break
                
                # Only proceed if we have both Ticket_Number and Travel_Sector columns
                if ticket_excel_col and sector_excel_col:
                    # Build keys using Ticket_Number value as PNR_Number
                    row_keys_ticket_as_pnr = {}
                    rows_with_valid_keys_ticket_as_pnr = []
                    
                    for idx in list(unmatched_indices):
                        row = batch_df.loc[idx]
                        
                        # Get Ticket_Number value
                        if ticket_excel_col not in row.index:
                            continue
                        
                        ticket_value = row[ticket_excel_col]
                        if self.is_empty_value(ticket_value):
                            continue
                        
                        # Get sector value
                        if sector_excel_col not in row.index:
                            continue
                        
                        sector_value = row[sector_excel_col]
                        if self.is_empty_value(sector_value):
                            continue
                        
                        # Split sector if it's multi-sector
                        sectors_to_query_ticket = self.split_multi_sector(sector_value)
                        
                        # Build keys for each sector using Ticket_Number as PNR_Number
                        key_values_list_ticket = []
                        for sector in sectors_to_query_ticket:
                            key_values_ticket = [ticket_value, sector]  # [PNR_Number, Travel_Sector]
                            key_values_list_ticket.append(tuple(key_values_ticket))
                        
                        if key_values_list_ticket:
                            row_keys_ticket_as_pnr[idx] = key_values_list_ticket
                            rows_with_valid_keys_ticket_as_pnr.append(idx)
                    
                    # If we have valid keys, execute batch query using PNR_Number + Travel_Sector
                    if row_keys_ticket_as_pnr:
                        # Collect all unique keys
                        all_keys_ticket_as_pnr = []
                        key_to_row_indices_ticket_as_pnr = {}
                        for idx, keys_list in row_keys_ticket_as_pnr.items():
                            for key in keys_list:
                                if key not in key_to_row_indices_ticket_as_pnr:
                                    all_keys_ticket_as_pnr.append(key)
                                    key_to_row_indices_ticket_as_pnr[key] = []
                                key_to_row_indices_ticket_as_pnr[key].append(idx)
                        
                        unique_keys_ticket_as_pnr = list(dict.fromkeys(all_keys_ticket_as_pnr))
                        
                        # Build batch query using PNR_Number and Travel_Sector
                        pnr_sector_combination = ['PNR_Number', 'Travel_Sector']
                        ref_cols_str_ticket = ', '.join([f"`{c}`" for c in pnr_sector_combination])
                        missing_columns_str_ticket = ', '.join([f"`{col}`" for col in missing_columns])
                        placeholders_ticket = ', '.join(["(" + ", ".join(["%s"] * len(pnr_sector_combination)) + ")" for _ in unique_keys_ticket_as_pnr])
                        
                        # Add NULL checks
                        null_checks_ticket = ' AND '.join([f"`{c}` IS NOT NULL" for c in pnr_sector_combination])
                        
                        query_ticket = (
                            f"SELECT {ref_cols_str_ticket}, {missing_columns_str_ticket} "
                            f"FROM `{table_name}` "
                            f"WHERE ({ref_cols_str_ticket}) IN ({placeholders_ticket}) "
                            f"AND {null_checks_ticket}"
                        )
                        params_ticket = [v for key in unique_keys_ticket_as_pnr for v in key]
                        
                        # Execute batch query
                        results_ticket_as_pnr = self.execute_query_with_retry(query_ticket, params_ticket)
                        
                        # Build lookup dictionary from results
                        lookup_ticket_as_pnr = {}
                        for r in results_ticket_as_pnr:
                            key = tuple(r[c] for c in pnr_sector_combination)
                            if key not in lookup_ticket_as_pnr:
                                lookup_ticket_as_pnr[key] = {col: r.get(col) for col in missing_columns}
                        
                        # Match rows with results
                        for idx in rows_with_valid_keys_ticket_as_pnr:
                            if idx in unmatched_indices:
                                keys_list = row_keys_ticket_as_pnr[idx]
                                
                                # Collect matching results
                                all_results_ticket = []
                                for key in keys_list:
                                    if key in lookup_ticket_as_pnr:
                                        all_results_ticket.append(lookup_ticket_as_pnr[key])
                                
                                if all_results_ticket:
                                    # Aggregate numeric columns
                                    aggregated_data_ticket = {}
                                    for col in missing_columns:
                                        if col in numeric_columns_to_aggregate:
                                            total = 0
                                            for result in all_results_ticket:
                                                val = result.get(col)
                                                if val is not None:
                                                    try:
                                                        total += float(val)
                                                    except (ValueError, TypeError):
                                                        pass
                                            aggregated_data_ticket[col] = total if total != 0 else None
                                        else:
                                            # For non-numeric columns, take first non-null value
                                            for result in all_results_ticket:
                                                val = result.get(col)
                                                if val is not None:
                                                    aggregated_data_ticket[col] = val
                                                    break
                                            if col not in aggregated_data_ticket:
                                                aggregated_data_ticket[col] = None
                                    
                                    row_matches[idx] = {
                                        'matched_data': aggregated_data_ticket,
                                        'combination': ['PNR_Number', 'Travel_Sector']  # Mark as PNR_Number match
                                    }
                                    unmatched_indices.remove(idx)
                                    combination_usage_stats["Ticket_Number_as_PNR_Number (fallback)"] = combination_usage_stats.get("Ticket_Number_as_PNR_Number (fallback)", 0) + 1
                                    logger.debug(f"Matched row {idx} using Ticket_Number value as PNR_Number: {ticket_value}")
            
            # Build enriched rows in original order (preserve exact row sequence)
            # Iterate by position to ensure order is maintained
            for pos in range(len(batch_df)):
                idx = batch_df.index[pos]
                row_data = batch_df.iloc[pos].to_dict()
                
                # Create a new row dict with all required columns
                complete_row = {}
                
                # Get matched data from database (if row was matched)
                matched_db_data = {}
                if idx in row_matches:
                    match_info = row_matches[idx]
                    matched_db_data = match_info['matched_data']
                    match_count += 1
                else:
                    no_match_count += 1
                
                # Process each column according to output_column_mapping
                for db_col, output_col, excel_only in output_column_mapping:
                    if excel_only:
                        # Excel-only columns: only take from Excel if column exists
                        excel_col_found = None
                        excel_names = excel_column_mappings.get(output_col, [])
                        for excel_name in excel_names:
                            matched_col = self.find_column_case_insensitive(excel_name, list(df_excel.columns))
                            if matched_col:
                                excel_col_found = matched_col
                                break
                        
                        if excel_col_found and excel_col_found in row_data:
                            # For Excel-only columns, use output_col as the key directly
                            complete_row[output_col] = row_data[excel_col_found]
                        else:
                            complete_row[output_col] = None
                    else:
                        # DB columns: always fetch from DB (even if exists in Excel)
                        # Use matched data from database
                        complete_row[db_col] = matched_db_data.get(db_col)
                
                enriched_data.append(complete_row)
        
        # Log combination usage statistics
        logger.info("Combination usage statistics:")
        for combo_str, count in combination_usage_stats.items():
            logger.info(f"  {combo_str}: {count} matches")
        
        # Create final DataFrame
        df_enriched = pd.DataFrame(enriched_data)
        
        # Build column rename mapping: internal column name -> output column name
        rename_dict = {}
        for db_col, output_col, excel_only in output_column_mapping:
            if excel_only:
                # Excel-only columns: output_col is already the key
                if output_col in df_enriched.columns:
                    rename_dict[output_col] = output_col  # Already correct name
            else:
                # DB columns: db_col -> output_col
                if db_col in df_enriched.columns:
                    rename_dict[db_col] = output_col
        
        # Apply column rename mapping
        if rename_dict:
            df_enriched = df_enriched.rename(columns=rename_dict)
        
        # Build final column order based on output_column_mapping
        final_column_order = []
        for db_col, output_col, excel_only in output_column_mapping:
            if output_col in df_enriched.columns:
                final_column_order.append(output_col)
        
        # Only include columns that exist in the DataFrame
        final_column_order = [col for col in final_column_order if col in df_enriched.columns]
        
        # Add any remaining columns that weren't in the mapping (shouldn't happen, but safety check)
        remaining_cols = [col for col in df_enriched.columns if col not in final_column_order]
        if remaining_cols:
            logger.warning(f"Found unexpected columns: {remaining_cols}")
            final_column_order.extend(remaining_cols)
        
        # Reorder DataFrame to match exact order
        df_enriched = df_enriched[final_column_order]
        
        logger.info(f"Sheet processing complete: {len(df_enriched)} rows, {match_count} matches, {no_match_count} no matches")
        logger.info(f"Output columns ({len(final_column_order)}): {final_column_order}")
        
        return df_enriched

    def enrich_data(self, excel_path: str, table_name: str, 
                   possible_reference_combinations: List[List[str]] = None,
                   column_mapping: Dict[str, str] = None,
                   output_path: Optional[str] = None):
        """
        Enhanced data enrichment with dynamic column detection and batch processing.
        Returns either pd.DataFrame (single sheet) or Dict[str, pd.DataFrame] (multiple sheets).
        """
        if column_mapping is None:
            column_mapping = {}
        if possible_reference_combinations is None:
            possible_reference_combinations = POSSIBLE_REFERENCE_COMBINATIONS
        
        # Validate file
        if not self.validate_file(excel_path):
            return None
        
        # Read file
        data = self.read_file_safely(excel_path)
        if data is None:
            return None
        
        # Handle multiple sheets
        if isinstance(data, dict):
            logger.info(f"Processing {len(data)} separate sheets")
            enriched_sheets = {}
            for sheet_name, df_sheet in data.items():
                logger.info(f"Processing sheet: {sheet_name}")
                df_enriched = self._enrich_single_dataframe(
                    df_sheet, table_name, possible_reference_combinations, column_mapping
                )
                if df_enriched is not None:
                    enriched_sheets[sheet_name] = df_enriched
            
            # Save output with separate sheets
            if output_path and enriched_sheets:
                try:
                    output_extension = os.path.splitext(output_path)[1].lower()
                    if output_extension == '.csv':
                        # For CSV, combine all sheets
                        df_combined = pd.concat(list(enriched_sheets.values()), ignore_index=True)
                        df_combined.to_csv(output_path, index=False)
                        logger.info(f"Data saved to: {output_path}")
                        
                        # Split data into credit_note and Invoice files based on Vendor K3 Amount
                        try:
                            logger.info("Splitting data into credit_note and Invoice files...")
                            credit_note_path, invoice_path = split_by_vendor_k3_amount(
                                output_path,
                                output_directory=os.path.dirname(output_path)
                            )
                            if credit_note_path:
                                logger.info(f"Credit note file created: {credit_note_path}")
                            if invoice_path:
                                logger.info(f"Invoice file created: {invoice_path}")
                        except Exception as split_error:
                            logger.warning(f"Could not split data into credit_note and Invoice files: {split_error}")
                    else:
                        # Save each sheet separately in Excel
                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            for sheet_name, df_enriched in enriched_sheets.items():
                                df_enriched.to_excel(writer, sheet_name=sheet_name, index=False)
                        logger.info(f"Data saved to {output_path} with {len(enriched_sheets)} sheets")
                        
                        # Apply header formatting
                        try:
                            self.apply_header_formatting(excel_path, output_path)
                        except Exception as format_error:
                            logger.warning(f"Could not apply header formatting: {format_error}")
                        
                        # Apply date formatting
                        try:
                            self.format_date_columns(output_path)
                        except Exception as date_format_error:
                            logger.warning(f"Could not apply date formatting: {date_format_error}")
                        
                        # Split data into credit_note and Invoice files based on Vendor K3 Amount
                        try:
                            logger.info("Splitting data into credit_note and Invoice files...")
                            credit_note_path, invoice_path = split_by_vendor_k3_amount(
                                output_path,
                                output_directory=os.path.dirname(output_path)
                            )
                            if credit_note_path:
                                logger.info(f"Credit note file created: {credit_note_path}")
                            if invoice_path:
                                logger.info(f"Invoice file created: {invoice_path}")
                        except Exception as split_error:
                            logger.warning(f"Could not split data into credit_note and Invoice files: {split_error}")
                except Exception as e:
                    logger.error(f"Error saving file: {e}")
            
            return enriched_sheets if enriched_sheets else None
        
        # Handle single sheet/DataFrame
        else:
            df_enriched = self._enrich_single_dataframe(
                data, table_name, possible_reference_combinations, column_mapping
            )
            
            # Save output
            if output_path and df_enriched is not None:
                try:
                    output_extension = os.path.splitext(output_path)[1].lower()
                    if output_extension == '.csv':
                        df_enriched.to_csv(output_path, index=False)
                        logger.info(f"Data saved to: {output_path}")
                        
                        # Split data into credit_note and Invoice files based on Vendor K3 Amount
                        try:
                            logger.info("Splitting data into credit_note and Invoice files...")
                            credit_note_path, invoice_path = split_by_vendor_k3_amount(
                                output_path,
                                output_directory=os.path.dirname(output_path)
                            )
                            if credit_note_path:
                                logger.info(f"Credit note file created: {credit_note_path}")
                            if invoice_path:
                                logger.info(f"Invoice file created: {invoice_path}")
                        except Exception as split_error:
                            logger.warning(f"Could not split data into credit_note and Invoice files: {split_error}")
                    else:
                        df_enriched.to_excel(output_path, index=False, engine='openpyxl')
                        logger.info(f"Data saved to: {output_path}")
                        
                        try:
                            self.apply_header_formatting(excel_path, output_path)
                        except Exception as format_error:
                            logger.warning(f"Could not apply header formatting: {format_error}")
                        
                        # Apply date formatting
                        try:
                            self.format_date_columns(output_path)
                        except Exception as date_format_error:
                            logger.warning(f"Could not apply date formatting: {date_format_error}")
                        
                        # Split data into credit_note and Invoice files based on Vendor K3 Amount
                        try:
                            logger.info("Splitting data into credit_note and Invoice files...")
                            credit_note_path, invoice_path = split_by_vendor_k3_amount(
                                output_path,
                                output_directory=os.path.dirname(output_path)
                            )
                            if credit_note_path:
                                logger.info(f"Credit note file created: {credit_note_path}")
                            if invoice_path:
                                logger.info(f"Invoice file created: {invoice_path}")
                        except Exception as split_error:
                            logger.warning(f"Could not split data into credit_note and Invoice files: {split_error}")
                except Exception as e:
                    logger.error(f"Error saving file: {e}")
            
            return df_enriched


class SFTPDownloader:
    """Simple SFTP client for downloading files from remote server."""
    
    def __init__(self, host: str, port: int, username: str, password: str):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.ssh_client = None
        self.sftp_client = None
    
    def connect(self) -> bool:
        try:
            self.ssh_client = paramiko.SSHClient()
            self.ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.ssh_client.connect(
                hostname=self.host,
                port=self.port,
                username=self.username,
                password=self.password,
                timeout=30
            )
            self.sftp_client = self.ssh_client.open_sftp()
            logger.info(f"Connected to SFTP server: {self.host}")
            return True
        except Exception as e:
            logger.error(f"Failed to connect to SFTP server: {e}")
            return False
    
    def disconnect(self):
        try:
            if self.sftp_client:
                self.sftp_client.close()
            if self.ssh_client:
                self.ssh_client.close()
            logger.info("SFTP connection closed")
        except Exception as e:
            logger.error(f"Error closing SFTP connection: {e}")
    
    def download_file(self, remote_path: str, local_dir: str) -> Optional[str]:
        try:
            os.makedirs(local_dir, exist_ok=True)
            filename = os.path.basename(remote_path)
            local_path = os.path.join(local_dir, filename)
            logger.info(f"Downloading {remote_path} to {local_path}")
            self.sftp_client.get(remote_path, local_path)
            logger.info(f"File downloaded successfully: {local_path}")
            return local_path
        except Exception as e:
            logger.error(f"Failed to download file: {e}")
            return None



class EmailSender:
    """Handles sending email notifications after processing."""
    
    def __init__(self, config: Dict):
        self.config = config
        self.enabled = config.get("enabled", False)
        self.recipient = config.get("recipient_email", "")
        self.smtp_server = config.get("smtp_server", "smtp.gmail.com")
        self.smtp_port = config.get("smtp_port", 587)
        self.sender_email = config.get("sender_email", "")
        self.sender_password = config.get("sender_password", "")
        self.subject = config.get("subject", "Data Merge Processing Report")
    
    def send_email(self, processing_result: Dict, log_file_path: Optional[str] = None, output_files: Optional[List[str]] = None) -> bool:
        """Send email notification with processing results."""
        if not self.enabled:
            logger.info("Email notifications are disabled")
            return False
        
        if not self.recipient or not self.sender_email:
            logger.warning("Email configuration incomplete - skipping email send")
            return False
        
        try:
            # Create email message
            msg = MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = self.recipient
            msg['Subject'] = self.subject
            
            # Create email body
            body = self._create_email_body(processing_result)
            msg.attach(MIMEText(body, 'html'))
            
            # Attach processed output files
            if output_files:
                for file_path in output_files:
                    if file_path and os.path.exists(file_path):
                        try:
                            with open(file_path, 'rb') as attachment:
                                part = MIMEBase('application', 'octet-stream')
                                part.set_payload(attachment.read())
                            encoders.encode_base64(part)
                            
                            # Determine MIME type based on file extension
                            file_ext = os.path.splitext(file_path)[1].lower()
                            if file_ext in ['.xlsx', '.xls', '.xlsb']:
                                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            elif file_ext == '.csv':
                                mime_type = 'text/csv'
                            else:
                                mime_type = 'application/octet-stream'
                            
                            part.add_header('Content-Type', mime_type)
                            part.add_header(
                                'Content-Disposition',
                                f'attachment; filename="{os.path.basename(file_path)}"'
                            )
                            msg.attach(part)
                            logger.info(f"Attached processed file: {os.path.basename(file_path)}")
                        except Exception as e:
                            logger.warning(f"Could not attach file {file_path}: {e}")
            
            # Attach log file if available
            if log_file_path and os.path.exists(log_file_path):
                try:
                    with open(log_file_path, 'rb') as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Type', 'text/plain')
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename="{os.path.basename(log_file_path)}"'
                    )
                    msg.attach(part)
                    logger.info(f"Attached log file: {os.path.basename(log_file_path)}")
                except Exception as e:
                    logger.warning(f"Could not attach log file: {e}")
            
            # Send email
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)
            
            logger.info(f"Email sent successfully to {self.recipient}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
            return False
    
    def _create_email_body(self, result: Dict) -> str:
        """Create HTML email body with processing results."""
        status = result.get("status", "unknown")
        processed = result.get("processed", 0)
        errors = result.get("errors", 0)
        results = result.get("results", [])
        
        # Determine status color
        if status == "completed" and errors == 0:
            status_color = "green"
        elif status == "completed" and errors > 0:
            status_color = "orange"
        else:
            status_color = "red"
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #4CAF50; color: white; padding: 10px; }}
                .content {{ padding: 20px; }}
                .status {{ color: {status_color}; font-weight: bold; }}
                .summary {{ background-color: #f5f5f5; padding: 15px; margin: 10px 0; border-radius: 5px; }}
                .file-list {{ margin-top: 15px; }}
                .file-item {{ padding: 5px; margin: 5px 0; }}
                .success {{ color: green; }}
                .error {{ color: red; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>Data Merge Processing Report</h2>
            </div>
            <div class="content">
                <p><strong>Processing Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><strong>Status:</strong> <span class="status">{status.upper()}</span></p>
                
                <div class="summary">
                    <h3>Summary</h3>
                    <p><strong>Files Processed Successfully:</strong> {processed}</p>
                    <p><strong>Errors:</strong> {errors}</p>
                    <p><strong>Total Files:</strong> {len(results)}</p>
                </div>
        """
        
        if results:
            html += """
                <div class="file-list">
                    <h3>File Details</h3>
            """
            for res in results:
                file_name = os.path.basename(res.get("file", "Unknown"))
                file_status = res.get("status", "unknown")
                status_class = "success" if file_status == "success" else "error"
                
                html += f"""
                    <div class="file-item">
                        <strong>{file_name}</strong> - <span class="{status_class}">{file_status.upper()}</span>
                """
                
                if file_status == "success":
                    rows = res.get("rows", 0)
                    output = res.get("output", "")
                    html += f"<br>Rows processed: {rows}"
                    if output:
                        html += f"<br>Output: {os.path.basename(output)}"
                
                if file_status in ["failed", "error"]:
                    error_msg = res.get("error", "Unknown error")
                    html += f"<br>Error: {error_msg}"
                
                html += "</div>"
            
            html += "</div>"
        
        html += """
            </div>
        </body>
        </html>
        """
        
        return html


class AutomatedProcessor:
    """Handles automated processing with scheduling."""
    
    def __init__(self, db_config: Dict, table_name: str, column_mapping: Dict, 
                 possible_reference_combinations: List[List[str]]):
        self.db_config = db_config
        self.table_name = table_name
        self.column_mapping = column_mapping
        self.possible_reference_combinations = possible_reference_combinations
        self.file_processor = FileProcessor(INPUT_DIRECTORY, OUTPUT_DIRECTORY, SUPPORTED_EXTENSIONS)
        self.is_running = False
        self.email_sender = EmailSender(EMAIL_CONFIG)
    
    def process_all_files(self) -> Dict[str, any]:
        """Process all files in the input directory."""
        logger.info("Starting automated file processing...")
        
        # Optional SFTP prefetch before discovering files
        try:
            if SFTP_CONFIG.get("enabled"):
                sftp_local_dir = SFTP_CONFIG.get("local_download_dir", INPUT_DIRECTORY)
                remote_path = SFTP_CONFIG.get("remote_file_path")
                if remote_path:
                    logger.info("SFTP prefetch enabled - attempting download from remote_file_path")
                    sftp = SFTPDownloader(
                        host=SFTP_CONFIG.get("host"),
                        port=SFTP_CONFIG.get("port", 22),
                        username=SFTP_CONFIG.get("username"),
                        password=SFTP_CONFIG.get("password")
                    )
                    if sftp.connect():
                        try:
                            downloaded = sftp.download_file(remote_path=remote_path, local_dir=sftp_local_dir)
                            if downloaded:
                                logger.info(f"SFTP file available at: {downloaded}")
                            else:
                                logger.warning("SFTP download did not produce a file")
                        finally:
                            sftp.disconnect()
                    else:
                        logger.error("Skipping SFTP download due to connection failure")
                else:
                    logger.info("SFTP enabled but no 'remote_file_path' provided; skipping download")
        except Exception as e:
            logger.error(f"SFTP prefetch error: {e}")
        
        # Discover files
        files_to_process = self.file_processor.discover_files()
        
        if not files_to_process:
            logger.info("No files found to process")
            return {"status": "no_files", "processed": 0, "errors": 0}
        
        # Initialize enricher
        enricher = DataEnricher(**self.db_config, debug_mode=DEBUG_MODE, debug_id=DEBUG_ID)
        
        processed_count = 0
        error_count = 0
        results = []
        
        try:
            # Connect to database
            if not enricher.connect():
                logger.error("Failed to connect to database")
                return {"status": "db_error", "processed": 0, "errors": len(files_to_process)}
            
            # Process each file
            for file_path in files_to_process:
                try:
                    logger.info(f"Processing file: {file_path}")
                    
                    # Generate output path
                    output_path = self.file_processor.get_output_path(file_path)
                    
                    # Enrich data
                    df_result = enricher.enrich_data(
                        excel_path=file_path,
                        table_name=self.table_name,
                        possible_reference_combinations=self.possible_reference_combinations,
                        column_mapping=self.column_mapping,
                        output_path=output_path
                    )
                    
                    if df_result is not None:
                        processed_count += 1
                        logger.info(f"Successfully processed: {file_path}")
                        
                        # Note: Original file is kept in place, not moved
                        
                        # Handle both DataFrame and dict (multiple sheets)
                        if isinstance(df_result, dict):
                            total_rows = sum(len(df) for df in df_result.values())
                            sheets_info = {name: len(df) for name, df in df_result.items()}
                            results.append({
                                "file": file_path,
                                "status": "success",
                                "rows": total_rows,
                                "sheets": len(df_result),
                                "sheets_info": sheets_info,
                                "output": output_path
                            })
                        else:
                            results.append({
                                "file": file_path,
                                "status": "success",
                                "rows": len(df_result),
                                "output": output_path
                            })
                    else:
                        error_count += 1
                        logger.error(f"Failed to process: {file_path}")
                        results.append({
                            "file": file_path,
                            "status": "failed",
                            "error": "Processing failed"
                        })
                        
                except Exception as e:
                    error_count += 1
                    logger.error(f"Error processing {file_path}: {e}")
                    results.append({
                        "file": file_path,
                        "status": "error",
                        "error": str(e)
                    })
            
        except Exception as e:
            logger.error(f"Critical error during processing: {e}")
            return {"status": "critical_error", "processed": processed_count, "errors": error_count}
        
        finally:
            enricher.disconnect()
        
        # Log summary
        logger.info(f"Processing complete: {processed_count} successful, {error_count} errors")
        
        result = {
            "status": "completed",
            "processed": processed_count,
            "errors": error_count,
            "results": results
        }
        
        # Send email notification
        if EMAIL_CONFIG.get("enabled", False):
            log_file_path = f"data_merge_{datetime.now().strftime('%Y%m%d')}.log"
            # Collect all output file paths from successful processing
            output_files = [res.get("output") for res in results if res.get("status") == "success" and res.get("output")]
            self.email_sender.send_email(result, log_file_path, output_files)
        
        return result
    
    def run_scheduled_job(self):
        """Run the scheduled processing job."""
        if self.is_running:
            logger.warning("Previous job still running, skipping this execution")
            return
        
        self.is_running = True
        try:
            logger.info("Starting scheduled processing job...")
            result = self.process_all_files()
            logger.info(f"Scheduled job completed: {result}")
            
            # Email is already sent in process_all_files if enabled
        except Exception as e:
            logger.error(f"Scheduled job failed: {e}")
            # Send error notification email
            if EMAIL_CONFIG.get("enabled", False):
                error_result = {
                    "status": "error",
                    "processed": 0,
                    "errors": 1,
                    "results": [{"file": "Scheduled Job", "status": "error", "error": str(e)}]
                }
                log_file_path = f"data_merge_{datetime.now().strftime('%Y%m%d')}.log"
                self.email_sender.send_email(error_result, log_file_path, None)
        finally:
            self.is_running = False
    
    def get_schedule_config(self):
        """Get scheduling configuration from loaded config."""
        schedule_config = CONFIG.get("scheduling", {})
        schedule_time = schedule_config.get("time", "13:00")
        enabled = schedule_config.get("enabled", True)
        return schedule_time, enabled
    
    def start_scheduler(self):
        """Start the scheduler using time from config.json."""
        schedule_time, enabled = self.get_schedule_config()
        
        if not enabled:
            logger.info("Scheduling is disabled in config.json")
            return
        
        logger.info(f"Setting up daily scheduler for {schedule_time} execution (from config.json)")
        schedule.every().day.at(schedule_time).do(self.run_scheduled_job)
        
        logger.info("Scheduler started. Waiting for scheduled execution...")
        while True:
            schedule.run_pending()
            time.sleep(1)  # Check every second for precise scheduling


# ====================================================================
# MAIN EXECUTION
# ====================================================================

if __name__ == "__main__":
    import sys
    
    # Check command line arguments
    if len(sys.argv) > 1:
        mode = sys.argv[1].lower()
    else:
        mode = "manual"  # Default mode
    
    print("="*60)
    print("ENHANCED DATA ENRICHMENT TOOL")
    print("="*60)
    print(f"Mode: {mode.upper()}")
    print(f"Input Directory: {INPUT_DIRECTORY}")
    print(f"Output Directory: {OUTPUT_DIRECTORY}")
    print(f"Database: {DB_CONFIG['database']}")
    print(f"Table: {TABLE_NAME}")
    print(f"Batch size: {BATCH_SIZE}")
    print(f"Debug mode: {DEBUG_MODE}")
    if SFTP_CONFIG.get("enabled"):
        print(f"SFTP: ON -> {SFTP_CONFIG.get('host')} | Remote: {SFTP_CONFIG.get('remote_file_path')} | Local: {SFTP_CONFIG.get('local_download_dir', INPUT_DIRECTORY)}")
    else:
        print("SFTP: OFF")
    print("="*60)
    
    # Initialize processor
    processor = AutomatedProcessor(
        db_config=DB_CONFIG,
        table_name=TABLE_NAME,
        column_mapping=COLUMN_MAPPING,
        possible_reference_combinations=POSSIBLE_REFERENCE_COMBINATIONS
    )
    
    if mode == "auto" or mode == "scheduler":
        # Run in automated/scheduled mode
        logger.info("Starting in automated mode with daily scheduling")
        try:
            processor.start_scheduler()
        except KeyboardInterrupt:
            logger.info("Scheduler stopped by user")
        except Exception as e:
            logger.error(f"Scheduler error: {e}")
    
    elif mode == "process":
        # Process all files once
        logger.info("Starting one-time processing of all files")
        try:
            result = processor.process_all_files()
            print("\n" + "="*60)
            print("PROCESSING COMPLETE!")
            print("="*60)
            print(f"Status: {result['status']}")
            print(f"Files processed: {result['processed']}")
            print(f"Errors: {result['errors']}")
            if 'results' in result:
                print("\nDetailed Results:")
                for res in result['results']:
                    print(f"  {res['file']}: {res['status']}")
            print("="*60)
        except Exception as e:
            logger.error(f"Processing error: {e}")
            print(f"\nERROR: {e}")
    
    else:
        # Manual mode - process single file (legacy behavior)
        logger.info("Starting in manual mode")
        
        # Check if input directory has files
        files = processor.file_processor.discover_files()
        if files:
            print(f"\nFound {len(files)} files in input directory:")
            for i, file in enumerate(files, 1):
                print(f"  {i}. {os.path.basename(file)}")
            
            if len(files) == 1:
                # Process the single file
                file_to_process = files[0]
                output_path = processor.file_processor.get_output_path(file_to_process)
                
                print(f"\nProcessing: {os.path.basename(file_to_process)}")
                print(f"Output: {output_path}")
                
                try:
                    result = processor.process_all_files()
                    print("\n" + "="*60)
                    print("SUCCESS!")
                    print("="*60)
                    print(f"Files processed: {result['processed']}")
                    print(f"Errors: {result['errors']}")
                    print("="*60)
                except Exception as e:
                    print(f"\nERROR: {e}")
                    logger.error(f"Manual processing error: {e}")
            else:
                print(f"\nMultiple files found. Use 'python data_merge.py process' to process all files")
                print("Or use 'python data_merge.py auto' to start automated processing")
        else:
            print(f"\nNo files found in {INPUT_DIRECTORY}")
            print("Please add Excel/CSV files to the input directory")